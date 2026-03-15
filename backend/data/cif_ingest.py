"""
Cities in Fiction (CIF) data ingestion.

Parses the ContributionsCIF.xlsx spreadsheet and geocodes the place names
to produce LiteraryPlace-compatible entries. These are high-quality, human-
curated records — mostly South Asian literature in regional languages.

Usage:
  python -m backend.data.cif_ingest
  python -m backend.data.cif_ingest --xlsx path/to/file.xlsx
"""

import asyncio
import json
import logging
import re
import sys
from pathlib import Path

import openpyxl
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut, GeocoderServiceError

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ── Pre-known coordinates for common cities ────────────────────────
# Avoids hitting the geocoder for well-known places.

KNOWN_COORDS: dict[str, list[float]] = {
    "Delhi": [77.2090, 28.6139],
    "New Delhi": [77.2090, 28.6139],
    "Mumbai": [72.8777, 19.0760],
    "Bombay": [72.8777, 19.0760],
    "Kolkata": [88.3639, 22.5726],
    "Calcutta": [88.3639, 22.5726],
    "Chennai": [80.2707, 13.0827],
    "Madras": [80.2707, 13.0827],
    "Bangalore": [77.5946, 12.9716],
    "Bengaluru": [77.5946, 12.9716],
    "Hyderabad": [78.4867, 17.3850],
    "Jaipur": [75.7873, 26.9124],
    "Lucknow": [80.9462, 26.8467],
    "Varanasi": [83.0007, 25.3176],
    "Goa": [74.1240, 15.2993],
    "Kochi": [76.2673, 9.9312],
    "Pune": [73.8567, 18.5204],
    "Ahmedabad": [72.5714, 23.0225],
    "Amritsar": [74.8723, 31.6340],
    "Shimla": [77.1734, 31.1048],
    "Mysore": [76.6394, 12.2958],
    "Lahore": [74.3587, 31.5204],
    "Karachi": [67.0011, 24.8607],
    "Islamabad": [73.0479, 33.6844],
    "Dhaka": [90.4125, 23.8103],
    "Colombo": [79.8612, 6.9271],
    "Kathmandu": [85.3240, 27.7172],
    "Kabul": [69.1723, 34.5281],
    "Patna": [85.1376, 25.6093],
    "Srinagar": [74.7973, 34.0837],
    "Darjeeling": [88.2631, 27.0360],
    "Mussoorie": [78.0644, 30.4598],
    "Aligarh": [78.0880, 27.8974],
    "Gorakhpur": [83.3732, 26.7606],
    "Shillong": [91.8933, 25.5788],
    "Jodhpur": [73.0243, 26.2389],
    "Kalimpong": [88.4700, 27.0660],
    "Jaisalmer": [70.9083, 26.9157],
    "Agra": [78.0081, 27.1767],
    "Bhopal": [77.4126, 23.2599],
    "Chandigarh": [76.7794, 30.7333],
    "Coimbatore": [76.9558, 11.0168],
    "Dehradun": [78.0322, 30.3165],
    "Indore": [75.8577, 22.7196],
    "Kanpur": [80.3319, 26.4499],
    "Madurai": [78.1198, 9.9252],
    "Mangalore": [74.8560, 12.9141],
    "Nagpur": [79.0882, 21.1458],
    "Pondicherry": [79.8083, 11.9416],
    "Thiruvananthapuram": [76.9366, 8.5241],
    "Udaipur": [73.7125, 24.5854],
    "Vizag": [83.2185, 17.6868],
    "Visakhapatnam": [83.2185, 17.6868],
    "Andhra Pradesh": [79.7400, 15.9129],
    "Mukteshwar": [79.6500, 29.4700],
    "Dhanbad": [86.4300, 23.7957],
    "Barasat": [88.4800, 22.7235],
    "Bhubaneswar": [85.8245, 20.2961],
    "Puri": [85.8315, 19.8135],
    "Banaras": [83.0007, 25.3176],
    "Aizawl": [92.7176, 23.7271],
    "Dajeeling": [88.2631, 27.0360],
    "Kumaon": [79.6500, 29.6000],
    "Dharamshala": [76.3234, 32.2190],
    "Kozhikode": [75.7804, 11.2588],
    "Mysuru": [76.6394, 12.2958],
    "Pasighat": [95.3269, 28.0670],
    "Belgavi": [74.4977, 15.8497],
    "Belgaum": [74.4977, 15.8497],
    "Casablanca": [-7.5898, 33.5731],
    "Ranaghat": [88.5700, 23.1800],
    "Srikakulam": [83.8961, 18.2949],
    "Kattupatti": [77.5700, 10.3200],
    "Kerala": [76.2711, 10.8505],
    "Tamil Nadu": [78.6569, 11.1271],
    "Karnataka": [75.7139, 15.3173],
    "Assam": [92.9376, 26.2006],
    "West Bengal": [87.8550, 22.9868],
    "Bengal": [87.8550, 22.9868],
    "Nepal": [84.1240, 28.3949],
    "North Sentinel Island": [92.2333, 11.5500],
    "Mahe": [75.5354, 11.7005],
    "Karikkottakkari": [75.3700, 11.8700],
    "Mandavgarh Fort": [75.3990, 22.2050],
    "Mandavgarh": [75.3990, 22.2050],
    "Bangarwadi": [74.4800, 18.9600],
    "Palasgaon": [75.9200, 19.5400],
    "Kija": [94.2600, 25.5800],
    "Gangauli": [83.5500, 25.8500],
    "Atharighat": [92.7800, 26.7500],
}

# ── City name normalization ────────────────────────────────────────

CITY_ALIASES = {
    "Bombay": "Mumbai",
    "Calcutta": "Kolkata",
    "Kolkata/ Calcutta": "Kolkata",
    "Kolkata/Calcutta": "Kolkata",
    "Madras": "Chennai",
    "Bengaluru": "Bangalore",
    "New Delhi": "Delhi",
    "Vizag": "Visakhapatnam",
}


def normalize_city(raw: str) -> tuple[str, str]:
    """
    Clean and normalize a city string.
    Returns (display_name, geocode_query).
    """
    raw = raw.strip()

    for alias, canonical in CITY_ALIASES.items():
        if alias.lower() in raw.lower():
            return canonical, canonical

    # Handle "City, State" patterns
    if "," in raw:
        parts = [p.strip() for p in raw.split(",")]
        return parts[0], raw

    # Handle "City/City" patterns — take the first
    if "/" in raw:
        parts = [p.strip() for p in raw.split("/")]
        return parts[0], parts[0]

    # Handle long descriptive locations
    if len(raw) > 40:
        # Try to extract a city name from parentheses
        paren = re.search(r'\(.*?(\w+)\)', raw)
        if paren:
            return paren.group(1), paren.group(1)
        words = raw.split()[:3]
        return " ".join(words), raw

    return raw, raw


def _geocode_sync(query: str) -> list[float] | None:
    """Geocode a place name using Nominatim."""
    geolocator = Nominatim(user_agent="Akhand-LitGeo/0.1", timeout=5)
    import time

    try:
        location = geolocator.geocode(query + ", India")
        if location:
            return [location.longitude, location.latitude]

        location = geolocator.geocode(query)
        if location:
            return [location.longitude, location.latitude]

        return None
    except (GeocoderTimedOut, GeocoderServiceError) as e:
        logger.warning(f"Geocoding failed for '{query}': {e}")
        return None


def geocode_city(raw_city: str) -> tuple[str, list[float] | None]:
    """
    Resolve a city name to coordinates.
    Uses pre-known coords first, then falls back to Nominatim.
    """
    display_name, geocode_query = normalize_city(raw_city)

    # Check pre-known coordinates
    for name in [display_name, geocode_query, raw_city]:
        if name in KNOWN_COORDS:
            return display_name, KNOWN_COORDS[name]

    # Partial match against known coords
    for known, coords in KNOWN_COORDS.items():
        if known.lower() in raw_city.lower():
            return display_name, coords

    # Fall back to Nominatim
    logger.info(f"Geocoding: '{geocode_query}' (from '{raw_city}')")
    coords = _geocode_sync(geocode_query)
    if coords:
        KNOWN_COORDS[display_name] = coords
        return display_name, coords

    return display_name, None


def _normalize_language(raw: str | None) -> str:
    if not raw:
        return "English"
    raw = raw.strip().rstrip(")")

    lang_map = {
        "english": "English",
        "hindi": "Hindi",
        "urdu": "Urdu",
        "bengali": "Bengali",
        "bangla": "Bengali",
        "tamil": "Tamil",
        "telugu": "Telugu",
        "marathi": "Marathi",
        "kannada": "Kannada",
        "malayalam": "Malayalam",
        "malaylam": "Malayalam",
        "odia": "Odia",
        "assamese": "Assamese",
        "konkani": "Konkani",
        "french": "French",
        "rajasthani": "Rajasthani",
    }

    for key, val in lang_map.items():
        if key in raw.lower():
            return val

    return raw


def _classify_form(form: str | None) -> list[str]:
    """Convert CIF form field to genre tags."""
    if not form:
        return ["literary fiction"]

    form_lower = form.strip().lower()
    mapping = {
        "novel": ["literary fiction"],
        "novella": ["literary fiction", "novella"],
        "short stor": ["short stories"],
        "poem": ["poetry"],
        "graphic novel": ["graphic novel"],
        "memoir": ["memoir"],
        "essay": ["essay"],
        "historical fiction": ["historical fiction"],
        "anthology": ["anthology"],
        "stories": ["short stories"],
    }

    for key, genres in mapping.items():
        if key in form_lower:
            return genres

    return ["literary fiction"]


def parse_cif_xlsx(xlsx_path: str | Path) -> list[dict]:
    """
    Parse the ContributionsCIF.xlsx spreadsheet.

    Columns (0-indexed from row 2):
      1: Title
      2: Author | Translator
      3: City
      4: Form
      5: Theme (actually a description/summary)
      6: Language
      7: Year of publication
      8: Comments
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    entries = []
    geocode_failures = []
    import time

    for row in range(3, ws.max_row + 1):
        title = ws.cell(row=row, column=1).value
        if not title or not str(title).strip():
            continue

        title = str(title).strip()
        author_raw = str(ws.cell(row=row, column=2).value or "").strip()
        city_raw = str(ws.cell(row=row, column=3).value or "").strip()
        form = str(ws.cell(row=row, column=4).value or "").strip()
        theme = str(ws.cell(row=row, column=5).value or "").strip()
        language = str(ws.cell(row=row, column=6).value or "").strip()
        year_raw = ws.cell(row=row, column=7).value
        comments = str(ws.cell(row=row, column=8).value or "").strip()

        if not city_raw or city_raw == "None":
            continue

        # Parse author (handle "Author; Translated by Translator" patterns)
        author = author_raw.split(";")[0].strip()
        translator = None
        if "translat" in author_raw.lower():
            parts = re.split(r";\s*(?:translated|trans\.)\s*(?:by|into)?\s*", author_raw, flags=re.IGNORECASE)
            if len(parts) >= 2:
                author = parts[0].strip()
                translator = parts[1].strip()

        # Parse year
        year = None
        if year_raw:
            try:
                year = int(float(str(year_raw)))
            except (ValueError, TypeError):
                pass

        # Geocode
        display_city, coords = geocode_city(city_raw)
        if not coords:
            geocode_failures.append(city_raw)
            continue

        # Build passage from theme + comments
        passage = theme if theme and theme != "None" else ""
        if comments and comments != "None":
            if passage:
                passage += " "
            passage += comments

        entry = {
            "id": f"cif-{row}-{display_city.lower().replace(' ', '-')}",
            "bookTitle": title,
            "author": author,
            "publishYear": year or 0,
            "placeName": display_city,
            "coordinates": coords,
            "placeType": "real",
            "settingType": "primary",
            "narrativeEra": f"{year}s" if year else "",
            "passage": passage,
            "sentiment": {
                "polarity": 0.0,
                "dominantEmotions": [],
                "themes": [],
            },
            "language": _normalize_language(language),
            "genres": _classify_form(form),
            "region": "South Asia",
            "source": "citiesinfiction",
        }

        if translator:
            entry["translator"] = translator

        entries.append(entry)
        time.sleep(0.05)  # rate limit for geocoder

    if geocode_failures:
        logger.warning(f"Failed to geocode {len(geocode_failures)} places: {set(geocode_failures)}")

    logger.info(f"Parsed {len(entries)} entries from CIF spreadsheet")
    return entries


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Ingest Cities in Fiction data")
    parser.add_argument("--xlsx", default="ContributionsCIF.xlsx", help="Path to CIF Excel file")
    parser.add_argument("--output", default=None, help="Output JSON path")
    parser.add_argument("--merge", action="store_true", help="Merge with existing Open Library data")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        logger.error(f"File not found: {xlsx_path}")
        sys.exit(1)

    entries = parse_cif_xlsx(xlsx_path)

    output_dir = Path(__file__).parent / "generated"
    output_dir.mkdir(parents=True, exist_ok=True)

    if args.merge:
        ol_path = output_dir / "literary_places.json"
        if ol_path.exists():
            with open(ol_path) as f:
                existing = json.load(f)
            ol_places = existing.get("places", [])
            logger.info(f"Merging with {len(ol_places)} existing Open Library entries")

            # Deduplicate by title + city
            seen = set()
            for p in ol_places:
                key = f"{p['bookTitle'].lower()}|{p['placeName'].lower()}"
                seen.add(key)

            new_entries = []
            for e in entries:
                key = f"{e['bookTitle'].lower()}|{e['placeName'].lower()}"
                if key not in seen:
                    new_entries.append(e)
                    seen.add(key)

            all_places = ol_places + new_entries
            logger.info(f"Added {len(new_entries)} new CIF entries (skipped {len(entries) - len(new_entries)} duplicates)")
        else:
            all_places = entries
    else:
        all_places = entries

    output_path = Path(args.output) if args.output else output_dir / "literary_places.json"

    with open(output_path, "w") as f:
        json.dump(
            {
                "version": "0.3.0",
                "sources": ["openlibrary", "citiesinfiction"] if args.merge else ["citiesinfiction"],
                "total": len(all_places),
                "places": all_places,
            },
            f,
            indent=2,
            ensure_ascii=False,
        )

    logger.info(f"Wrote {len(all_places)} places to {output_path}")

    from collections import Counter
    cities = Counter(p["placeName"] for p in all_places)
    langs = Counter(p["language"] for p in all_places)

    print(f"\n{'='*60}")
    print(f"CIF INGESTION: {len(all_places)} literary places")
    print(f"{'='*60}")
    print(f"\nTop cities:")
    for city, count in cities.most_common(15):
        print(f"  {city}: {count}")
    print(f"\nLanguages:")
    for lang, count in langs.most_common():
        print(f"  {lang}: {count}")
    print(f"\nOutput: {output_path}")


if __name__ == "__main__":
    main()
